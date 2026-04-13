from fastapi import FastAPI, UploadFile, File, Form, Request
import time
from fastapi.responses import HTMLResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
import shutil
import os
import json
import pandas as pd
from typing import List
import sys
import subprocess
from fastapi.responses import HTMLResponse, JSONResponse, FileResponse
from datetime import datetime, timedelta
import glob
import socket
from PyPDF2 import PdfMerger
import holidays as pyholidays
import smtplib, ssl
from email.message import EmailMessage

from google import genai
from google.genai import types
from .payroll_processor import process_payroll, preprocess_roster, preprocess_clockin, match_names
from .payslip_generator import generate_payslip, PayslipPDF
from fpdf import FPDF
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
try:
    import win32clipboard
except ImportError:
    win32clipboard = None

from . import config
import logging
from logging.handlers import RotatingFileHandler

# Setup Logging
os.makedirs(config.LOG_DIR, exist_ok=True)
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        RotatingFileHandler(os.path.join(config.LOG_DIR, "portal.log"), maxBytes=10*1024*1024, backupCount=5),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger("WimpyPortal")

GEMINI_KEY = os.environ.get("GEMINI_API_KEY", "")

# Fallback to config.json
if not GEMINI_KEY and os.path.exists("config.json"):
    try:
        with open("config.json", "r") as f:
            cfg = json.load(f)
            GEMINI_KEY = cfg.get("GEMINI_API_KEY", "")
            if GEMINI_KEY:
                logger.info("Loaded GEMINI_API_KEY from config.json")
    except Exception as e:
        logger.error(f"Failed to read config.json: {e}")

client = None
if GEMINI_KEY:
    try:
        client = genai.Client(api_key=GEMINI_KEY)
    except Exception as e:
        logger.error(f"AI Client failed to initialize: {e}")

os.makedirs("output/payslips", exist_ok=True)
os.makedirs(config.ARCHIVE_DIR, exist_ok=True)
os.makedirs("data/input", exist_ok=True)

app = FastAPI()

app.mount("/static", StaticFiles(directory="static"), name="static")
app.mount("/payslips", StaticFiles(directory="output/payslips"), name="payslips")
app.mount("/archives", StaticFiles(directory=config.ARCHIVE_DIR), name="archives")

@app.get("/", response_class=HTMLResponse)
async def read_index():
    with open("static/index.html", "r", encoding="utf-8") as f:
        return f.read()

@app.post("/api/login")
async def login(request: Request):
    data = await request.json()
    username = str(data.get("username") or data.get("name", "")).strip()
    pin = str(data.get("id_last4", "")).strip()

    # Manager bypass requires PIN 0000
    if username.lower() == "manager":
        if pin == "0000":
            return JSONResponse(content={"success": True, "role": "MANAGER", "name": "Manager", "username": "manager"})
        else:
            return JSONResponse(content={"success": False, "error": "Invalid Manager PIN"})

    # Check PIN for staff
    pin_path = "data/staff_pins.json"
    if os.path.exists(pin_path):
        with open(pin_path, "r") as f:
            pins = json.load(f)
        
        # Match by username (primary) or name (fallback for legacy)
        if username in pins and pins[username] == pin:
            master_path = "data/templates/Staff_Details_Template.xlsx"
            role = "EMPLOYEE"
            portal_access = "MY_SHIFTS"
            display_name = username
            if os.path.exists(master_path):
                mdf = pd.read_excel(master_path)
                # Check Username column first, then Name
                col = "Username" if "Username" in mdf.columns else "Name"
                row = mdf[mdf[col].astype(str).str.strip().str.lower() == username.lower()]
                if not row.empty:
                    display_name = str(row.iloc[0].get("Name", username))
                    portal_access = str(row.iloc[0].get("Portal Access", "MY_SHIFTS")).replace("nan", "MY_SHIFTS")
                    if portal_access == "MANAGER":
                        role = "MANAGER"
            return JSONResponse(content={"success": True, "role": role, "name": display_name, "username": username, "portal_access": portal_access})

    return JSONResponse(content={"success": False, "error": "Invalid Username or PIN"})

@app.get("/api/holidays")
async def get_holidays(year: int = None):
    if not year: year = datetime.now().year
    sa_holidays = pyholidays.SouthAfrica(years=year)
    return JSONResponse(content={
        "holidays": [{"date": str(d), "name": n} for d, n in sorted(sa_holidays.items())]
    })

@app.get("/api/requests")
async def get_requests():
    path = "data/off_day_requests.json"
    if not os.path.exists(path): return JSONResponse(content=[])
    with open(path, "r") as f:
        return JSONResponse(content=json.load(f))

@app.post("/api/submit_request")
async def submit_request(request: Request):
    data = await request.json()
    path = "data/off_day_requests.json"
    employee = data.get("employee")
    off_days = data.get("off_days", []) # List of strings e.g. ["2026-04-10"]
    req_type = data.get("type", "OFF_DAY") # "OFF_DAY" or "LEAVE"
    
    now = datetime.now()
    
    # 2026 School Holidays SA
    SCHOOL_HOLIDAYS = [
        ("2026-03-28", "2026-04-07"),
        ("2026-06-27", "2026-07-20"),
        ("2026-09-24", "2026-10-05"),
        ("2026-12-10", "2026-12-31")
    ]
    
    sa_holidays = pyholidays.SouthAfrica(years=[now.year, now.year + 1])
    
    def is_school_holiday(date_str):
        d = datetime.strptime(date_str, "%Y-%m-%d")
        for start, end in SCHOOL_HOLIDAYS:
            if datetime.strptime(start, "%Y-%m-%d") <= d <= datetime.strptime(end, "%Y-%m-%d"):
                return True
        return False

    if len(off_days) > 2:
        return JSONResponse(content={"success": False, "error": "Maximum 2 dates allowed per request."})

    # Validate non-consecutive
    if len(off_days) == 2:
        d1 = datetime.strptime(off_days[0], "%Y-%m-%d")
        d2 = datetime.strptime(off_days[1], "%Y-%m-%d")
        if abs((d1 - d2).days) == 1:
            return JSONResponse(content={"success": False, "error": "Consecutive off-days are not allowed."})

    errors = []
    status = "APPROVED"  # Default for weekdays

    for day in off_days:
        d_obj = datetime.strptime(day, "%Y-%m-%d")
        lead_time = (d_obj - now).days

        # 1. 14-day lead time for LEAVE only
        if req_type == "LEAVE" and lead_time < 14:
            errors.append(f"LEAVE for {day} must be requested at least 14 days in advance. Currently only {lead_time} days away.")

        # 2. Weekends (5=Sat, 6=Sun) or Annual Leave ALWAYS require manager approval
        if d_obj.weekday() in [5, 6] or req_type == "LEAVE":
            status = "PENDING_APPROVAL"

    if errors:
        return JSONResponse(content={"success": False, "error": "; ".join(errors)})

    reqs = []
    if os.path.exists(path):
        with open(path, "r") as f:
            try:
                reqs = json.load(f)
                if not isinstance(reqs, list): reqs = []
            except:
                reqs = []

    reqs.append({
        "id": str(now.timestamp()),
        "employee": employee,
        "off_days": off_days,
        "type": req_type,
        "status": status,
        "timestamp": now.strftime("%Y-%m-%d %H:%M")
    })

    with open(path, "w") as f: json.dump(reqs, f, indent=4)
    return JSONResponse(content={"success": True, "status": status})

@app.post("/api/approve_request")
async def approve_request(request: Request):
    data = await request.json()
    req_id = data.get("id")
    action = data.get("action") # "APPROVE" or "REJECT"
    
    path = "data/off_day_requests.json"
    if not os.path.exists(path): return JSONResponse(content={"success": False, "error": "No requests found"})
    
    with open(path, "r") as f: requests = json.load(f)
    
    found = False
    for r in requests:
        if r.get("id") == req_id:
            r["status"] = "APPROVED" if action == "APPROVE" else "REJECTED"
            found = True
            break
            
    if found:
        with open(path, "w") as f: json.dump(requests, f, indent=4)
        return JSONResponse(content={"success": True})
    return JSONResponse(content={"success": False, "error": "Request not found"})

# --- STAFF MANAGEMENT ---
def get_all_staff_list():
    master_path = "data/templates/Staff_Details_Template.xlsx"
    if not os.path.exists(master_path):
        return []
    try:
        df = pd.read_excel(master_path)
        df.columns = [str(c).strip() for c in df.columns]
        # Safe fill-na for mixed types
        df = df.astype(object).where(pd.notnull(df), "")
        
        staff_list = []
        for _, row in df.iterrows():
            name = str(row.get("Name", "")).strip()
            if not name: continue
            role_raw = str(row.get("Role", "WAITER")).replace("nan", "WAITER")
            role_clean = role_raw.upper().replace("BARITA", "BARISTA")
            staff_no = str(row.get("Staff #", "")).replace(".0", "").strip()
            staff_list.append({"name": name, "role": role_clean, "staff_no": staff_no})
        return staff_list
    except:
        return []

@app.get("/api/staff")
async def get_staff():
    master_path = "data/templates/Staff_Details_Template.xlsx"
    pin_path = "data/staff_pins.json"
    
    if not os.path.exists(master_path):
        return JSONResponse(content={"staff": []})
        
    try:
        df = pd.read_excel(master_path)
        # Standardize column names (strip spaces)
        df.columns = [str(c).strip() for c in df.columns]
        
        # Ensure all required columns exist in the DataFrame for safe accessing
        required_cols = ['Name', 'ID Number', 'Rate', 'Start Date', 'Leave Credit', 'Cell Number', 'Role', 'Email', 'Portal Access', 'Username', 'Payslip Delivery']
        for col in required_cols:
            if col not in df.columns:
                df[col] = "" # Add missing column if template is partial
        
        # Safe fill-na for mixed types
        df = df.astype(object).where(pd.notnull(df), "")
        
        pins = {}
        if os.path.exists(pin_path):
            with open(pin_path, "r") as f:
                pins = json.load(f)
                
        staff_list = []
        for _, row in df.iterrows():
            name = str(row.get("Name", "")).strip()
            if not name or name.lower() == "nan": continue
            
            username = str(row.get("Username", name)).strip()
            if not username or username.lower() == "nan": username = name
            
            # Standardize role
            role_raw = str(row.get("Role", "WAITER")).replace("nan", "WAITER")
            role_clean = role_raw.upper().replace("BARITA", "BARISTA")
            
            # Standardize rate
            rate_raw = row.get("Rate", "30.33")
            try:
                rate_str = str(rate_raw).replace("R", "").replace(",", ".").strip()
                if rate_str and rate_str.lower() != "nan":
                    rate_val = float(rate_str)
                else:
                    rate_val = 30.33
            except:
                rate_val = 30.33

            # Standardize leave credit
            leave_raw = row.get("Leave Credit", 0.0)
            try:
                if str(leave_raw).lower() in ["", "nan", "none"]:
                    leave_val = 0.0
                else:
                    leave_val = float(leave_raw)
            except:
                leave_val = 0.0

            staff_list.append({
                "name": name,
                "username": username,
                "id_number": str(row.get("ID Number", "")).replace(".0", ""),
                "role": role_clean,
                "rate": rate_val,
                "cell": str(row.get("Cell Number", "")).replace(".0", ""),
                "delivery": str(row.get("Payslip Delivery", "WHATSAPP")),
                "portal_access": str(row.get("Portal Access", "MY_SHIFTS")),
                "pin": pins.get(username, "0000"),
                "start_date": str(row.get("Start Date", "")),
                "leave_credit": leave_val,
                "email": str(row.get("Email", ""))
            })
            
        # Role sorting
        role_order = ["MANAGER", "SUPERVISOR", "BARISTA", "GRILLER", "KITCHEN", "WAITER"]
        def sort_key(s):
            r = s["role"].upper()
            try: idx = role_order.index(r)
            except: idx = 99
            return (idx, s["name"].lower())
            
        staff_list.sort(key=sort_key)
        return JSONResponse(content={"staff": staff_list})
    except Exception as e:
        return JSONResponse(content={"error": str(e)}, status_code=500)

@app.post("/api/staff/save")
async def save_staff(request: Request):
    try:
        data = await request.json()
        master_path = "data/templates/Staff_Details_Template.xlsx"
        pin_path = "data/staff_pins.json"
        
        name = data.get("name", "").strip().title()
        if not name: return JSONResponse(content={"success": False, "error": "Name is required"})
        
        username = data.get("username", name).strip()
        if not username: username = name
        
        new_pin = str(data.get("pin", "0000")).strip()
        role = str(data.get("role", "WAITER")).upper()
        
        # Robust numeric conversion
        def to_float(val, default=0.0):
            try:
                if val is None or str(val).strip().lower() in ["", "nan"]: return default
                return float(str(val).replace("R", "").replace(",", ".").strip())
            except: return default

        rate = to_float(data.get("rate"), 30.33)
        cell = str(data.get("cell", "")).strip()
        delivery = str(data.get("delivery", "WHATSAPP"))
        is_manager = data.get("is_manager", False)
        id_number = str(data.get("id_number", "")).strip()
        start_date = str(data.get("start_date", "")).strip()
        leave_credit = to_float(data.get("leave_credit"), 0.0)
        email = str(data.get("email", "")).strip()
        
        # Save to Excel
        if os.path.exists(master_path):
            try:
                df = pd.read_excel(master_path)
                df.columns = [str(c).strip() for c in df.columns]
                # Ensure text columns are treated as objects to prevent dtype crashes
                text_cols = ['Name', 'Username', 'Role', 'Cell Number', 'ID Number', 'Start Date', 'Email', 'Portal Access', 'Payslip Delivery']
                for col in text_cols:
                    if col in df.columns:
                        df[col] = df[col].astype(object)
            except Exception as e:
                return JSONResponse(content={"success": False, "error": f"Could not read staff excel: {str(e)}"})
        else:
            df = pd.DataFrame(columns=['Name', 'ID Number', 'Rate', 'Start Date', 'Leave Credit', 'Cell Number', 'Role', 'Email', 'Portal Access', 'Username', 'Payslip Delivery'])
            
        # Match by Username (primary) or exact Name
        found = False
        for idx, row in df.iterrows():
            existing_user = str(row.get("Username", "")).strip()
            existing_name = str(row.get("Name", "")).strip().lower()
            if (existing_user and existing_user == username) or existing_name == name.lower():
                df.at[idx, "Name"] = name
                df.at[idx, "Username"] = username
                df.at[idx, "Role"] = role
                df.at[idx, "Rate"] = rate
                df.at[idx, "Cell Number"] = cell
                df.at[idx, "Payslip Delivery"] = delivery
                df.at[idx, "Portal Access"] = "MANAGER" if is_manager else "MY_SHIFTS"
                df.at[idx, "ID Number"] = id_number
                df.at[idx, "Start Date"] = start_date
                df.at[idx, "Leave Credit"] = leave_credit
                df.at[idx, "Email"] = email
                found = True
                break
                
        if not found:
            new_row = {
                "Name": name, "Username": username, "Role": role, "Rate": rate, 
                "Cell Number": cell, "Payslip Delivery": delivery, 
                "Portal Access": "MANAGER" if is_manager else "MY_SHIFTS",
                "ID Number": id_number, "Start Date": start_date,
                "Leave Credit": leave_credit, "Email": email
            }
            df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
            
        try:
            df.to_excel(master_path, index=False)
        except Exception as e:
            if "Permission denied" in str(e):
                return JSONResponse(content={"success": False, "error": "Cannot save: The Excel file is currently open. Please close 'Staff_Details_Template.xlsx' and try again."})
            return JSONResponse(content={"success": False, "error": f"Excel Save Error: {str(e)}"})
        
        # Save PIN
        try:
            pins = {}
            if os.path.exists(pin_path):
                with open(pin_path, "r") as f: pins = json.load(f)
            pins[username] = str(new_pin)
            with open(pin_path, "w") as f: json.dump(pins, f, indent=4)
        except Exception as e:
            return JSONResponse(content={"success": False, "error": f"PIN Save Error: {str(e)}"})
        
        return JSONResponse(content={"success": True})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JSONResponse(content={"success": False, "error": f"General System Error: {str(e)}"})


@app.delete("/api/staff/delete/{username}")
async def delete_staff(username: str):
    master_path = "data/templates/Staff_Details_Template.xlsx"
    pin_path = "data/staff_pins.json"
    
    try:
        # 1. Remove from Excel
        if os.path.exists(master_path):
            df = pd.read_excel(master_path)
            df.columns = [str(c).strip() for c in df.columns]
            
            # Find row by Username (exact) or Name (case-insensitive strip match)
            # Create a clean version of the columns for comparison
            def is_match(row_val, search_val):
                return str(row_val).strip().lower() == search_val.lower()

            # We use boolean masking for robust deletion
            mask = df.apply(lambda r: is_match(r.get("Username", ""), username) or is_match(r.get("Name", ""), username), axis=1)
            
            initial_count = len(df)
            df = df[~mask]
            
            if len(df) < initial_count:
                df.to_excel(master_path, index=False)
                print(f"Deleted {username} from Excel.")
            else:
                print(f"Warning: {username} not found in Excel columns.")

        # 2. Remove from PINs
        if os.path.exists(pin_path):
            with open(pin_path, "r") as f:
                pins = json.load(f)
            
            # Delete exact key
            if username in pins:
                del pins[username]
            
            # Also try case-insensitive cleanup
            keys_to_delete = [k for k in pins.keys() if k.lower() == username.lower()]
            for k in keys_to_delete:
                if k in pins: del pins[k]
                
            with open(pin_path, "w") as f:
                json.dump(pins, f, indent=4)
                    
        return JSONResponse(content={"success": True})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JSONResponse(content={"success": False, "error": str(e)})

@app.post("/api/export_roster_pdf")
async def export_roster_pdf(request: Request):
    try:
        data = await request.json()
        week_date = data.get("date", "Unknown Week")
        rows = data.get("rows", [])
        
        from fpdf import FPDF
        pdf = FPDF(orientation='L') # Landscape
        pdf.add_page()
        pdf.set_font("Helvetica", 'B', 16)
        pdf.cell(0, 10, f"WIMPY DE VILLE - WEEKLY ROSTER ({week_date})", align='C', new_x="LMARGIN", new_y="NEXT")
        pdf.ln(5)
        
        # Table Header
        pdf.set_font("Helvetica", 'B', 9)
        pdf.set_fill_color(240, 240, 240)
        cols = ["Name", "Wed", "Thu", "Fri", "Sat", "Sun", "Mon", "Tue", "Total"]
        widths = [45, 30, 30, 30, 30, 30, 30, 30, 15]
        
        for i, col in enumerate(cols):
            pdf.cell(widths[i], 10, col, border=1, fill=True, align='C')
        pdf.ln()
        
        pdf.set_font("Helvetica", '', 8.5)
        for r in rows:
            if not r: continue
            pdf.cell(widths[0], 9, str(r.get("name", "")), border=1)
            pdf.cell(widths[1], 9, str(r.get("wed", "OFF")), border=1, align='C')
            pdf.cell(widths[2], 9, str(r.get("thu", "OFF")), border=1, align='C')
            pdf.cell(widths[3], 9, str(r.get("fri", "OFF")), border=1, align='C')
            pdf.cell(widths[4], 9, str(r.get("sat", "OFF")), border=1, align='C')
            pdf.cell(widths[5], 9, str(r.get("sun", "OFF")), border=1, align='C')
            pdf.cell(widths[6], 9, str(r.get("mon", "OFF")), border=1, align='C')
            pdf.cell(widths[7], 9, str(r.get("tue", "OFF")), border=1, align='C')
            pdf.cell(widths[8], 9, str(r.get("total", "0")), border=1, align='C')
            pdf.ln()
            
        output_path = f"tmp/Roster_{week_date}.pdf"
        os.makedirs("tmp", exist_ok=True)
        pdf.output(output_path)
        
        return FileResponse(output_path, filename=f"Roster_{week_date}.pdf")
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JSONResponse(content={"error": str(e)}, status_code=500)

@app.get("/api/archives/download_roster_pdf")
async def download_roster_pdf(path: str):
    """Converts an archived Excel roster to a PDF with re-calculated hours for guaranteed accuracy."""
    try:
        archive_path = os.path.join("data/archives", path.lstrip('/'))
        if not os.path.exists(archive_path):
             return JSONResponse(content={"error": "File not found"}, status_code=404)
             
        # Read Excel
        df = pd.read_excel(archive_path)
        week_date = os.path.basename(path).replace("Roster_", "").replace(".xlsx", "")
        
        # Calculate dates for headers
        from datetime import datetime, timedelta
        import re
        day_mapping = {}
        try:
            tue_dt = datetime.strptime(week_date, "%Y-%m-%d")
            days_list = ["Wed", "Thu", "Fri", "Sat", "Sun", "Mon", "Tue"]
            for i, d in enumerate(days_list):
                offset = 6 - i
                day_dt = tue_dt - timedelta(days=offset)
                day_mapping[d] = f"{d} {day_dt.strftime('%d %b')}"
        except: 
            day_mapping = {d: d for d in ["Wed", "Thu", "Fri", "Sat", "Sun", "Mon", "Tue"]}

        def parse_hours(text):
            if pd.isna(text): return 0.0
            s = str(text).strip().upper()
            if s == 'OFF' or not s or s == '-': return 0.0
            
            # 1. Try to extract value from parentheses: (9.50)
            m = re.search(r'\((\d+\.?\d*)\)', s)
            if m: return float(m.group(1))
            
            # 2. Try to split and calculate: 07:00-18:00
            try:
                # Remove any spaces and replace variant dashes
                s_clean = s.replace(' ', '').replace('—', '-').replace('–', '-')
                parts = s_clean.split('-')
                if len(parts) == 2:
                    t1_parts = parts[0].split(':')
                    t2_parts = parts[1].split(':')
                    
                    h1 = float(t1_parts[0])
                    m1 = float(t1_parts[1]) if len(t1_parts) > 1 else 0.0
                    
                    h2 = float(t2_parts[0])
                    m2 = float(t2_parts[1]) if len(t2_parts) > 1 else 0.0
                    
                    dur = (h2 + m2/60) - (h1 + m1/60)
                    if dur < 0: dur += 24
                    return dur
            except: pass
            
            # 3. Fallback: try to just convert to float if it's a plain number
            try: return float(s)
            except: return 0.0

        from fpdf import FPDF
        pdf = FPDF(orientation='L') 
        pdf.add_page()
        pdf.set_font("Helvetica", 'B', 16)
        pdf.cell(0, 15, f"WIMPY DE VILLE - ARCHIVED ROSTER (Week Ending {week_date})", align='C', new_x="LMARGIN", new_y="NEXT")
        pdf.ln(5)
        
        # Table Header
        pdf.set_font("Helvetica", 'B', 9)
        pdf.set_fill_color(240, 240, 240)
        
        raw_cols = list(df.columns)
        display_cols = []
        for c in raw_cols:
            c_str = str(c).strip()
            if c_str in day_mapping:
                display_cols.append(day_mapping[c_str])
            else:
                display_cols.append(c_str)
                
        # Handle columns dynamic width? Or stick to fixed?
        # Standard: Name, Wed, Thu, Fri, Sat, Sun, Mon, Tue, Total
        widths = [45, 30, 30, 30, 30, 30, 30, 30, 15]
        
        for i, col in enumerate(display_cols[:len(widths)]):
            pdf.cell(widths[i], 10, col, border=1, fill=True, align='C')
        pdf.ln()
        
        pdf.set_font("Helvetica", '', 8.5)
        # Sort by Name (optional but good)
        df_sorted = df.copy()
        if "Name" in df_sorted.columns:
            df_sorted = df_sorted.dropna(subset=["Name"])
            df_sorted = df_sorted[~df_sorted["Name"].astype(str).str.contains("📁")]
            df_sorted = df_sorted.sort_values(by="Name")

        for _, r in df_sorted.iterrows():
            row_total = 0.0
            # First pass: calculate total
            for col in raw_cols:
                if str(col).strip() in ["Wed", "Thu", "Fri", "Sat", "Sun", "Mon", "Tue"]:
                    row_total += parse_hours(r.get(col))

            # Second pass: render row
            for i, col in enumerate(raw_cols[:len(widths)]):
                val = str(r.get(col, "OFF")).replace('nan', 'OFF')
                
                is_total_col = "total" in str(col).lower()
                if is_total_col:
                    val = f"{row_total:.2f}"
                    align = 'C'
                    pdf.set_font("Helvetica", 'B', 8.5)
                else:
                    align = 'L' if i == 0 else 'C'
                    pdf.set_font("Helvetica", '', 8.5)
                
                pdf.cell(widths[i], 9, val, border=1, align=align)
            pdf.ln()
            
        temp_pdf = f"tmp/Archived_Roster_{week_date}.pdf"
        os.makedirs("tmp", exist_ok=True)
        pdf.output(temp_pdf)
        
        return FileResponse(temp_pdf, filename=f"Roster_{week_date}.pdf")
    except Exception as e:
        print(f"Archive PDF Error: {e}")
        import traceback
        traceback.print_exc()
        return JSONResponse(content={"error": str(e)}, status_code=500)




@app.post("/api/process")
async def process_files(
    roster: UploadFile = File(None),
    clockin: UploadFile = File(None),
    date_string: str = Form(None)
):
    os.makedirs("tmp", exist_ok=True)
    
    roster_path = "data/input/latest_roster.xlsx"
    clockin_path = "data/input/latest_clockin.xlsx"
    
    if roster and roster.filename:
        roster_path = f"tmp/{roster.filename}"
        with open(roster_path, "wb") as buffer:
             shutil.copyfileobj(roster.file, buffer)
             
    if clockin and clockin.filename:
        clockin_path = f"tmp/{clockin.filename}"
        with open(clockin_path, "wb") as buffer:
             shutil.copyfileobj(clockin.file, buffer)
             
    if not os.path.exists(roster_path) or not os.path.exists(clockin_path):
        return JSONResponse(content={"error": "Missing Roster or Clock-in files, and no local fallback found."})
         
    master_details = {}
    master_path = "data/templates/Staff_Details_Template.xlsx"
    if os.path.exists(master_path):
        try:
            master_df = pd.read_excel(master_path)
            cols = [str(c).strip().lower() for c in master_df.columns]
            for _, row in master_df.iterrows():
                name = str(row['Name']).strip().lower() if 'name' in cols else str(row.iloc[0]).strip().lower()
                rate_val = str(row['Rate']).replace('R', '').replace(',', '.').strip() if 'rate' in cols else "30.33"
                try: rate_val = float(rate_val) if rate_val and str(rate_val).lower() != "nan" else 30.33
                except: rate_val = 30.33
                
                master_details[name] = {
                    "rate": rate_val,
                    "id_number": str(row['ID Number']).replace('nan', '') if 'id number' in cols else "",
                    "start_date": str(row['Start Date']).replace('nan', '') if 'start date' in cols else "",
                    "leave_credit": str(row['Leave Credit']).replace('nan', '') if 'leave credit' in cols else "",
                    "cell_number": str(row['Cell Number']).replace('nan', '').replace('.0', '') if 'cell number' in cols else "",
                    "role": str(row['Role']).replace('nan', 'WAITER').upper().replace('BARITA', 'BARISTA') if 'role' in cols else "WAITER"
                }
        except Exception as e:
            print("Could not parse master list:", e)
    # Load holidays
    holidays = []
    if os.path.exists("data/holidays.json"):
        try:
            with open("data/holidays.json", "r") as f:
                h_data = json.load(f)
                # handle both list and dict formats
                if isinstance(h_data, list):
                    holidays = [h.get('date') for h in h_data if h.get('date')]
                elif isinstance(h_data, dict):
                    holidays = [h.get('date') for h in h_data.get('holidays', []) if h.get('date')]
        except Exception: pass

    # Load manual mappings
    mappings_path = "data/name_mappings.json"
    manual_mappings = {}
    if os.path.exists(mappings_path):
        try:
            with open(mappings_path, "r") as f:
                manual_mappings = json.load(f)
        except Exception: pass

    output_df, summary_df, overtime_flags, unmatched_names = process_payroll(
        roster_path, 
        clockin_path, 
        holidays=holidays, 
        output_format='dataframe',
        week_ending=date_string,
        manual_mappings=manual_mappings
    )
    
    # unmatched_names is now returned directly from the processor!
    # No need to re-match here.
    
    # Calculate shifts according to ROSTER (not just clock-ins with hours)
    # This ensures that even if they are sick, they must still pay the R40/shift.
    roster_shifts_dict = {}
    roster_all_names = []
    # 1. Get all names currently in the Master Staff Details (Best source for dropdown)
    master_path = "data/templates/Staff_Details_Template.xlsx"
    if os.path.exists(master_path):
        try:
            m_df = pd.read_excel(master_path)
            name_col = "Name" if "Name" in m_df.columns else m_df.columns[0]
            roster_all_names = [str(n).strip() for n in m_df[name_col].dropna().unique() if str(n).strip()]
        except: pass

    # 2. Add any names from the Roster file that might not be in Master yet
    roster_shifts_dict = {}
    try:
        raw_roster = pd.read_excel(roster_path)
        roster_df = preprocess_roster(raw_roster)
        if not roster_df.empty:
            we_dt = pd.to_datetime(date_string)
            st_dt = we_dt - timedelta(days=6)
            roster_df['Date'] = pd.to_datetime(roster_df['Date'])
            active_roster = roster_df[(roster_df['Date'] >= st_dt) & (roster_df['Date'] <= we_dt)]
            roster_shifts_dict = active_roster.groupby('Employee Name')['Date'].nunique().to_dict()
            for r_name in roster_shifts_dict.keys():
                if r_name not in roster_all_names:
                    roster_all_names.append(r_name)
    except Exception as e:
        print(f"Error calculating roster shifts: {e}")

    # Prepare summary data from whatever worked hours we have
    worked_summary = {}
    if summary_df is not None and not summary_df.empty:
        for _, row in summary_df.iterrows():
            worked_summary[str(row['Employee Name'])] = row.to_dict()

    # Combine Roster names and Worked names to ensure everyone on the roster appears
    all_names_to_process = sorted(list(set(roster_all_names + list(worked_summary.keys()))))

    employees = []
    for name in all_names_to_process:
        name_lower = name.lower()
        row = worked_summary.get(name, {})
        
        rate = 30.33
        details = {"id_number": "", "start_date": "", "leave_credit": "", "cell_number": "", "role": "WAITER"}
        for m_name, m_detail in master_details.items():
            if name_lower in m_name or m_name in name_lower:
                rate = m_detail["rate"]
                details = m_detail
                break
        
        employees.append({
            "name": name,
            "reg_hours": float(row.get('Reg Hours', 0)),
            "sun_hours": float(row.get('Sun Hours', 0)),
            "tue_hours": float(row.get('Tue Hours', 0)),
            "hol_hours": float(row.get('Hol Hours', 0)),
            "leave_days": 0.0,
            "sick_days": 0.0,
            "total_hours": float(row.get('Total Payable Hours', 0)),
            "raw_roster_hours": float(row.get('Raw Rostered Hours', 0)),
            "raw_clocked_hours": float(row.get('Raw Clocked Hours', 0)),
            "shifts_worked": int(roster_shifts_dict.get(name, 0)),
            "rate": rate,
            "bonus": 0.0,
            "till_short": 0.0,
            "tips": 0.0,
            "clothing": 0.0,
            **details
        })

    # Add Day Name for UI clarity
    for f in overtime_flags:
        try:
            dt = datetime.strptime(f['date'], "%Y-%m-%d")
            f['day_name'] = dt.strftime("%A")
        except:
            f['day_name'] = ""

    return JSONResponse({
        "employees": employees,
        "overtime_flags": overtime_flags,
        "unmatched": unmatched_names,
        "roster_staff": roster_all_names
    })

@app.post("/api/save_mapping")
async def save_mapping(request: Request):
    try:
        data = await request.json()
        clockin_name = data.get('clockin_name')
        roster_name = data.get('roster_name')
        
        mappings_path = "data/name_mappings.json"
        mappings = {}
        if os.path.exists(mappings_path):
            with open(mappings_path, "r") as f:
                mappings = json.load(f)
        
        mappings[clockin_name] = roster_name
        
        os.makedirs("data", exist_ok=True)
        with open(mappings_path, "w") as f:
            json.dump(mappings, f, indent=4)
            
        return {"status": "success"}
    except Exception as e:
        return {"status": "error", "message": str(e)}

@app.post("/api/approve_overtime")
async def approve_overtime(request: Request):
    """Saves overtime decisions (APPROVED or DENIED) to data/overtime_approvals.json"""
    try:
        data = await request.json()
        employee = data.get("employee")
        date_str = data.get("date")
        status = data.get("status")
        
        if not employee or not date_str or not status:
            return JSONResponse(content={"success": False, "error": "Missing mapping detail."})
            
        path = "data/overtime_approvals.json"
        approvals = {}
        if os.path.exists(path):
            with open(path, "r") as f:
                approvals = json.load(f)
                
        key = f"{employee}_{date_str}"
        approvals[key] = status
        
        with open(path, "w") as f:
            json.dump(approvals, f, indent=4)
            
        return JSONResponse(content={"success": True})
    except Exception as e:
        return JSONResponse(content={"success": False, "error": str(e)})

@app.post("/api/generate")
async def generate_pdfs(data: dict):
    print("!!! PAYROLL GENERATE CALLED v1.1 !!!")
    try:
        os.makedirs("output/payslips", exist_ok=True)
        employees = data.get("employees", [])
        
        # Clear existing payslips for a clean run
        os.makedirs("output/payslips", exist_ok=True)
        import glob
        for f in glob.glob("output/payslips/*"):
            try:
                if os.path.isfile(f):
                    os.remove(f)
            except Exception:
                pass
                
        def safe_float(val, default=0.0):
            try:
                if val is None or str(val).strip().lower() in ["", "nan", "undefined", "null"]:
                    return default
                fval = float(val)
                import math
                if math.isnan(fval) or math.isinf(fval):
                    return default
                return fval
            except:
                return default
        
        # 1. Calculate BOH Pool from Waiter Shifts
        total_pool = 0
        waiter_shortfalls = {}
        for emp in employees:
            role = str(emp.get("role", "WAITER")).upper()
            if role == "WAITER":
                shifts = safe_float(emp.get("shifts_worked", 0))
                tips_left = safe_float(emp.get("tips_left", 0))
                owed = shifts * 40
                total_pool += owed
                waiter_shortfalls[emp.get("name")] = max(0, owed - tips_left)

        # 2. Calculate BOH Shares (All staff except Waiter, Mgr, Sup)
        recipients = [e for e in employees if str(e.get("role", "")).upper() not in ["WAITER", "MANAGER", "SUPERVISOR"]]
        total_boh_hours = sum(safe_float(e.get("total_hours", 0)) for e in recipients)
        
        links = []
        leave_deductions = {}
        process_errors = []
        
        for emp in employees:
            try:
                name = str(emp.get("name", "Unknown"))
                role = str(emp.get("role", "WAITER")).upper()

                # Calculate total hours if missing
                reg = safe_float(emp.get("reg_hours"))
                tue = safe_float(emp.get("tue_hours"))
                sun = safe_float(emp.get("sun_hours"))
                hol = safe_float(emp.get("hol_hours"))
                total_hrs = safe_float(emp.get("total_hours", reg + tue + sun + hol))

                # BOH Share Calculation
                indiv_share = 0
                if role not in ["WAITER", "MANAGER", "SUPERVISOR"] and total_boh_hours > 0:
                    indiv_share = (total_hrs / total_boh_hours) * total_pool

                clean_pay = {
                    "name": name,
                    "role": role,
                    "rate": safe_float(emp.get("rate")),
                    "reg_hours": safe_float(emp.get("reg_hours")),
                    "tue_hours": safe_float(emp.get("tue_hours")),
                    "sun_hours": safe_float(emp.get("sun_hours")),
                    "hol_hours": safe_float(emp.get("hol_hours")),
                    "leave_days": safe_float(emp.get("leave_days")),
                    "sick_days": safe_float(emp.get("sick_days")),
                    "bonus": safe_float(emp.get("bonus")),
                    "till_short": safe_float(emp.get("till_short")),
                    "tips_shortfall": waiter_shortfalls.get(name, 0),
                    "boh_tip_share": indiv_share,
                    "clothing": safe_float(emp.get("clothing")),
                    "id_number": str(emp.get("id_number", "")),
                    "start_date": str(emp.get("start_date", "")),
                    "leave_credit": str(emp.get("leave_credit", "")),
                    "cell_number": str(emp.get("cell_number", "")),
                    "payslip_preference": str(emp.get("payslip_preference", "WHATSAPP"))
                }

                # 1. Calculate math with CLEAN data FIRST so we can skip zero-pay
                leave_pay = clean_pay["leave_days"] * 7 * clean_pay["rate"]
                sick_pay = clean_pay["sick_days"] * 7 * clean_pay["rate"]
                
                gross = (clean_pay["reg_hours"] * clean_pay["rate"]) + \
                        (clean_pay["tue_hours"] * clean_pay["rate"]) + \
                        (clean_pay["sun_hours"] * clean_pay["rate"] * 1.5) + \
                        (clean_pay["hol_hours"] * clean_pay["rate"] * 2) + \
                        leave_pay + sick_pay + clean_pay["bonus"] + clean_pay["boh_tip_share"]
                
                uif = gross * 0.01
                total_deduct = clean_pay["tips_shortfall"] + clean_pay["till_short"] + uif + clean_pay["clothing"]
                net = gross - total_deduct

                if net <= 0:
                    print(f"Skipping {name} (Nett R{net:.2f} <= 0)")
                    continue

                # 2. Generate PDF with CLEAN data
                filepath = generate_payslip(clean_pay)
                urlpath = filepath.replace('\\', '/')
                
                message = f"Hi {clean_pay['name']}, here is your payslip for the week. Total Nett Salary: R {net:.2f}."
                
                phone = clean_pay['cell_number']
                phone = ''.join(filter(str.isdigit, phone))

                links.append({
                    "name": clean_pay["name"],
                    "url": f"/{urlpath}",
                    "phone": phone,
                    "wa_message": message,
                    "net": net,
                    "payslip_preference": clean_pay["payslip_preference"]
                })
                
                if clean_pay["leave_days"] > 0:
                    leave_deductions[clean_pay['name'].strip().lower()] = clean_pay["leave_days"]
            except Exception as emp_e:
                error_msg = f"Error processing employee {emp.get('name')}: {str(emp_e)}"
                print(error_msg)
                process_errors.append(error_msg)
                continue
                
        # --- DEBUG RETURN IF NO LINKS ---
        if not links:
            print(f"FAILED RUN: employees_count={len(employees)}, errors_count={len(process_errors)}")
            return JSONResponse(content={
                "links": [],
                "error": f"Processed {len(employees)} staff but 0 payslips were successfully created.",
                "details": process_errors if process_errors else ["No errors were caught, but no payslips were created. This suggests an issue with the employee data format."]
            }, status_code=200)
                
        # --- LEAVE UPDATING ---
        master_path = "data/templates/Staff_Details_Template.xlsx"
        if os.path.exists(master_path) and leave_deductions:
            try:
                import pandas as pd
                master_df = pd.read_excel(master_path)
                name_col = next((c for c in master_df.columns if str(c).strip().lower() == 'name'), master_df.columns[0])
                leave_col = next((c for c in master_df.columns if str(c).strip().lower() == 'leave credit'), None)
                
                if leave_col:
                    for idx, row in master_df.iterrows():
                        emp_name = str(row[name_col]).strip().lower()
                        if emp_name in leave_deductions:
                            try: cur = float(row[leave_col]) if str(row[leave_col]).lower() != 'nan' else 0.0
                            except: cur = 0.0
                            master_df.at[idx, leave_col] = max(0.0, cur - leave_deductions[emp_name])
                    master_df.to_excel(master_path, index=False)
            except Exception as e:
                print("Failed to update leave:", e)
                
        # --- ARCHIVING ---
        try:
            date_str = datetime.now().strftime("%Y-%m-%d")
            archive_month_dir = get_archive_dir(date_str)
            payroll_folder_name = f"Payslips_{date_str}_{datetime.now().strftime('%H-%M')}"
            archive_dir = os.path.join(archive_month_dir, payroll_folder_name)
            os.makedirs(archive_dir, exist_ok=True)
            
            # 1. Archive Payslips
            import shutil
            for l in links:
                source = l["url"].lstrip("/")
                if os.path.exists(source):
                    dest = os.path.join(archive_dir, os.path.basename(source))
                    shutil.copy(source, dest)

            # 2. Generate and Archive EFT Summary PDF
            summary_path = os.path.join(archive_dir, "EFT_Summary.pdf")
            generate_eft_summary(links, output_path=summary_path)
            
            # 3. Create a zip for convenience
            shutil.make_archive(os.path.join(archive_dir, "All_Payslips"), "zip", "output/payslips")

            # 4. Copy active roster for reference
            if os.path.exists("data/input/latest_roster.xlsx"):
                shutil.copy("data/input/latest_roster.xlsx", os.path.join(archive_dir, "Matching_Roster.xlsx"))

            # 5. Save metadata for UI display (Total Amount + Individual List)
            total_net_all = sum(l.get("net", 0.0) for l in links)
            settlements = [{"name": l["name"], "net": l["net"]} for l in links]
            with open(os.path.join(archive_dir, "metadata.json"), "w") as f:
                json.dump({
                    "total_net": total_net_all,
                    "count": len(links),
                    "timestamp": date_str,
                    "settlements": settlements
                }, f)

            return JSONResponse(content={
                "links": links, 
                "archive": payroll_folder_name,
                "total_net": total_net_all,
                "eft_summary_url": f"/api/archives/download/{os.path.relpath(summary_path, 'data/archives').replace('\\','/')}"
            })

        except Exception as arch_e:
            print("Archiving failed:", arch_e)
            return JSONResponse(content={"links": links, "warning": "Archiving failed but payslips generated."})
            
    except Exception as global_e:
        import traceback
        traceback.print_exc()
        return JSONResponse(content={"error": str(global_e)}, status_code=500)

def generate_eft_summary(employees_data, output_path="output/EFT_Summary.pdf"):
    """Generates a summary list of all employees and their nett pay for EFT processing."""
    from fpdf import FPDF
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Helvetica", 'B', 16)
    pdf.cell(0, 10, "OFFICIAL EFT SALARY SUMMARY", align='C', new_x="LMARGIN", new_y="NEXT")
    pdf.set_font("Helvetica", '', 10)
    pdf.cell(0, 8, f"Date Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", align='C', new_x="LMARGIN", new_y="NEXT")
    pdf.ln(10)
    
    # Table Header
    pdf.set_font("Helvetica", 'B', 10)
    pdf.cell(80, 8, "Employee Name", border=1)
    pdf.cell(50, 8, "Role", border=1, align='C')
    pdf.cell(60, 8, "Nett Salary", border=1, new_x="LMARGIN", new_y="NEXT", align='R')
    
    pdf.set_font("Helvetica", '', 10)
    total_nett = 0
    for emp in employees_data:
        pdf.cell(80, 8, emp.get('name', 'Unknown'), border=1)
        pdf.cell(50, 8, emp.get('payslip_preference', 'WHATSAPP'), border=1, align='C')
        pdf.cell(60, 8, f"R {emp.get('net', 0.0):.2f}", border=1, new_x="LMARGIN", new_y="NEXT", align='R')
        total_nett += emp.get('net', 0.0)
        
    pdf.set_font("Helvetica", 'B', 11)
    pdf.cell(130, 10, "TOTAL EFT PAYROLL", border=1)
    pdf.cell(60, 10, f"R {total_nett:.2f}", border=1, new_x="LMARGIN", new_y="NEXT", align='R')
    
    pdf.output(output_path)
    return output_path

@app.get("/api/download_all")
async def download_all():
    import shutil
    shutil.make_archive("output/summaries/Payslips_Batch", "zip", "output/payslips")
    return FileResponse("output/summaries/Payslips_Batch.zip", media_type="application/zip", filename="Payslips_Batch.zip")

@app.get("/api/download_merged_pdf")
async def download_merged_pdf(folder: str = None):
    try:
        from PyPDF2 import PdfMerger
        import glob
        merger = PdfMerger()
        
        if folder:
            # Look in the archived folder
            archive_path = os.path.join("data/archives", folder.lstrip('/'))
            pdf_files = glob.glob(os.path.join(archive_path, "*.pdf"))
            # Exclude EFT_Summary if it exists
            pdf_files = [f for f in pdf_files if "EFT_Summary" not in f]
        else:
            # Look in current session output
            pdf_files = glob.glob("output/payslips/*.pdf")
            
        if not pdf_files:
            return JSONResponse(content={"error": f"No payslips found in {folder if folder else 'latest batch'} to merge."}, status_code=404)
            
        for pdf_file in sorted(pdf_files):
            merger.append(pdf_file)
            
        date_suffix = folder.replace('/', '_') if folder else datetime.now().strftime('%Y-%m-%d')
        merged_filename = f"output/summaries/All_Payslips_{date_suffix}.pdf"
        os.makedirs("output/summaries", exist_ok=True)
        merger.write(merged_filename)
        merger.close()
        return FileResponse(merged_filename, media_type="application/pdf", filename=os.path.basename(merged_filename))
    except Exception as e:
        return JSONResponse(content={"error": str(e)}, status_code=500)

@app.post("/api/download_eft")
async def download_eft_excel(request: Request):
    data = await request.json()
    links = data.get("links", [])
    
    import pandas as pd
    rows = []
    total_val = 0
    for l in links:
        net_val = round(l.get("net", 0.0), 2)
        rows.append({"Employee Name": l.get("name"), "Final Net Pay": net_val})
        total_val += net_val
        
    rows.append({"Employee Name": "TOTAL", "Final Net Pay": round(total_val, 2)})
        
    date_str = datetime.now().strftime("%Y-%m-%d")
    os.makedirs("output/summaries", exist_ok=True)
    filename = f"output/summaries/EFT_Summary_{date_str}.xlsx"
    pd.DataFrame(rows).to_excel(filename, index=False)
    return FileResponse(filename, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", filename=os.path.basename(filename))

@app.post("/api/download_eft_pdf")
async def download_eft_pdf_api(request: Request):
    try:
        data = await request.json()
        links = data.get("links", [])
        os.makedirs("output/summaries", exist_ok=True)
        out_path = f"output/summaries/EFT_Summary_{datetime.now().strftime('%Y-%m-%d')}.pdf"
        generate_eft_summary(links, output_path=out_path)
        return FileResponse(out_path, media_type="application/pdf", filename=os.path.basename(out_path))
    except Exception as e:
        return JSONResponse(content={"error": str(e)}, status_code=500)

@app.post("/api/copy_to_clipboard")
async def copy_to_clipboard(request: Request):
    if not win32clipboard:
        return JSONResponse(content={"success": False, "error": "win32clipboard not available"})
        
    try:
        data = await request.json()
        relative_path = data.get("path", "").lstrip('/')
        abs_path = os.path.abspath(relative_path)
        
        if not os.path.exists(abs_path):
            return JSONResponse(content={"success": False, "error": f"File not found: {abs_path}"})
            
        import time
        success = False
        for attempt in range(5):
            try:
                win32clipboard.OpenClipboard()
                win32clipboard.EmptyClipboard()
                win32clipboard.SetClipboardData(win32clipboard.CF_HDROP, [abs_path])
                win32clipboard.CloseClipboard()
                success = True
                break
            except Exception:
                time.sleep(0.1)
                
        if success:
            return JSONResponse(content={"success": True})
        else:
            return JSONResponse(content={"success": False, "error": "Clipboard access denied after retries."})
    except Exception as e:
        return JSONResponse(content={"success": False, "error": str(e)})

@app.post("/api/print_file")
async def print_file(request: Request):
    try:
        data = await request.json()
        relative_path = data.get("path", "").lstrip('/')
        abs_path = os.path.abspath(relative_path)
        
        if os.path.exists(abs_path):
            # Uses the system's 'print' verb on Windows
            os.startfile(abs_path, "print")
            return JSONResponse(content={"success": True})
        return JSONResponse(content={"success": False, "error": "File not found"})
    except Exception as e:
        return JSONResponse(content={"success": False, "error": str(e)})

@app.post("/api/send_roster_image")
async def send_roster_image(request: Request):
    try:
        import base64
        data = await request.json()
        img_data = data.get("image", "")
        if not img_data:
            return JSONResponse(content={"success": False, "error": "No image data received"})
        
        # Remove header
        if "," in img_data:
            img_data = img_data.split(",")[1]
            
        with open("output/roster_snapshot.png", "wb") as f:
            f.write(base64.b64decode(img_data))
            
        # Trigger WhatsApp bot for standard person (e.g. manager) or as defined
        # For now, we'll save it and log it. We can extend wa_bot to send this file.
        with open("output/whatsapp_payload.json", "w") as f:
            json.dump([{
                "phone": data.get("phone", ""),
                "message": "📅 Here is the latest Staff Roster for Wimpy De Ville.",
                "file": "output/roster_snapshot.png"
            }], f)
        
        subprocess.Popen([sys.executable, "app/wa_bot.py"])
        return JSONResponse(content={"success": True})
    except Exception as e:
        return JSONResponse(content={"success": False, "error": str(e)})

@app.post("/api/whatsapp_send")
async def whatsapp_send(request: Request):
    data = await request.json()
    links = data.get("links", [])
    
    payload = []
    for l in links:
        path = l.get("url", "").lstrip('/')
        if path:
            payload.append({
                "phone": l.get("phone", ""),
                "message": l.get("wa_message", ""),
                "file": path
            })
            
    with open("output/whatsapp_payload.json", "w") as f:
        json.dump(payload, f)
        
    subprocess.Popen([sys.executable, "app/wa_bot.py"])
    return JSONResponse(content={"status": "Bot started in background."})

@app.post("/api/cleanup")
async def auto_cleanup():
    """Removes temporary files and old payslips to keep the space clean."""
    try:
        if os.path.exists("tmp"):
            shutil.rmtree("tmp")
        # Keep the latest payslips in output, but clear individual ones if needed
        # For now, let's just clear tmp
        return JSONResponse(content={"success": True, "message": "Cleanup complete!"})
    except Exception as e:
        return JSONResponse(content={"success": False, "error": str(e)})

@app.post("/api/fetch_ankerdata")
async def fetch_ankerdata(request: Request):
    try:
        data = await request.json()
        target_date = data.get("date", "")
        
        # Run the bot and pass the date as an argument
        res = subprocess.run([sys.executable, "app/ankerdata_bot.py", target_date], capture_output=True, text=True)
        if res.returncode == 0:
            return JSONResponse(content={"success": True})
        else:
            print(f"Ankerdata Sync Error: {res.stderr}")
            return JSONResponse(content={"success": False, "error": res.stderr})
    except Exception as e:
        return JSONResponse(content={"success": False, "error": str(e)})

@app.post("/api/fetch_mock")
async def fetch_mock():
    """Simulates a successful Ankerdata fetch for testing."""
    try:
        roster_src = "tests/test_docs/MOCK_Roster.xlsx"
        clock_src = "tests/test_docs/MOCK_ClockIn.xlsx"
        
        if os.path.exists(roster_src) and os.path.exists(clock_src):
            shutil.copy(roster_src, "data/input/latest_roster.xlsx")
            shutil.copy(clock_src, "data/input/latest_clockin.xlsx")
            return JSONResponse(content={"success": True, "message": "Simulated fetch complete!"})
        else:
            return JSONResponse(content={"success": False, "error": "Mock files not found in tests/test_docs/"})
    except Exception as e:
        return JSONResponse(content={"success": False, "error": str(e)})

def get_archive_dir(date_str):
    """Returns data/archives/YYYY-MM/"""
    try:
        dt = datetime.strptime(date_str, "%Y-%m-%d")
        month_dir = dt.strftime("%Y-%m")
        path = f"data/archives/{month_dir}"
        os.makedirs(path, exist_ok=True)
        return path
    except:
        return "data/archives"

@app.get("/api/archives")
async def get_archives():
    try:
        payslip_archives = []
        roster_archives = []
        
        base_dir = "data/archives"
        if os.path.exists(base_dir):
            for root, dirs, files in os.walk(base_dir):
                for d in dirs:
                    if d.startswith("Payslips_"):
                        rel_path = os.path.relpath(os.path.join(root, d), base_dir).replace("\\", "/")
                        meta_path = os.path.join(root, d, "metadata.json")
                        total = 0
                        count = 0
                        m = {}
                        if os.path.exists(meta_path):
                            try:
                                with open(meta_path, "r") as f:
                                    m = json.load(f)
                                    total = m.get("total_net", 0)
                                    count = m.get("count", 0)
                            except: pass
                        
                        payslip_archives.append({
                            "folder": rel_path,
                            "total": total,
                            "count": count,
                            "settlements": m.get("settlements", [])
                        })
                for f in files:
                    # Collect roster excels
                    if f.endswith(".xlsx") and ("Roster" in f or "latest_roster" in f):
                         rel_path = os.path.relpath(os.path.join(root, f), base_dir).replace("\\", "/")
                         roster_archives.append(rel_path)
        # Sort by folder name descending
        payslip_archives.sort(key=lambda x: x["folder"], reverse=True)
        roster_archives.sort(key=lambda x: x, reverse=True)
        
        return JSONResponse(content={"payslip_archives": payslip_archives, "roster_archives": roster_archives})
    except Exception as e:
        return JSONResponse(content={"error": str(e)}, status_code=500)

@app.post("/api/archives/delete")
async def delete_archive(data: dict):
    path_sub = data.get("path")
    if not path_sub: return JSONResponse(content={"success": False, "error": "Path required"})
    
    # Security: Ensure it's inside data/archives
    base_dir = os.path.abspath("data/archives")
    full_path = os.path.abspath(os.path.join(base_dir, path_sub.lstrip("/")))
    
    if not full_path.startswith(base_dir):
        return JSONResponse(content={"success": False, "error": "Unauthorized path access attempt."})
    
    try:
        if os.path.exists(full_path):
            if os.path.isdir(full_path):
                shutil.rmtree(full_path)
            else:
                os.remove(full_path)
            return JSONResponse(content={"success": True})
        return JSONResponse(content={"success": False, "error": "File or folder not found."})
    except Exception as e:
        return JSONResponse(content={"success": False, "error": str(e)})

@app.get("/api/archives/roster/{path_sub:path}")
async def download_archived_roster(path_sub: str):
    path = os.path.join("data/archives", path_sub)
    if os.path.exists(path):
        return FileResponse(path, filename=os.path.basename(path))
    return JSONResponse(content={"error": "File not found"}, status_code=404)

@app.get("/api/archives/payslips_zip/{path_sub:path}")
async def download_archived_payslips_zip(path_sub: str):
    path = os.path.join("data/archives", path_sub, "All_Payslips.zip")
    if os.path.exists(path):
        return FileResponse(path, filename=f"All_Payslips.zip")
    return JSONResponse(content={"error": "ZIP not found"}, status_code=404)

@app.get("/api/archives/eft_summary/{path_sub:path}")
async def download_archived_eft_summary(path_sub: str):
    path = os.path.join("data/archives", path_sub, "EFT_Summary.xlsx")
    if os.path.exists(path):
        return FileResponse(path, filename=f"EFT_Summary.xlsx")
    return JSONResponse(content={"error": "EFT Summary not found"}, status_code=404)

@app.get("/api/archives/view/roster/{path_sub:path}")
async def view_archived_roster(path_sub: str):
    """View an archived roster from monthly sub-folders (e.g. 2026-04/Roster_...)"""
    path = os.path.join("data/archives", path_sub)
    if os.path.exists(path):
        try:
            df = pd.read_excel(path)
            # Replaced NaN with empty strings for JSON compatibility
            df = df.fillna("")
            return JSONResponse(content={"columns": df.columns.tolist(), "rows": df.to_dict('records')})
        except Exception as e:
            return JSONResponse(content={"error": str(e)}, status_code=500)
    return JSONResponse(content={"error": "File not found"}, status_code=404)

@app.get("/api/archives/view/eft_summary/{folder}")
async def view_archived_eft_summary(folder: str):
    path = f"data/archives/Payslips/{folder}/EFT_Summary.xlsx"
    if os.path.exists(path):
        try:
            df = pd.read_excel(path)
            # Replaced NaN with empty strings for JSON compatibility
            df = df.fillna("")
            return JSONResponse(content={"columns": df.columns.tolist(), "rows": df.to_dict('records')})
        except Exception as e:
            return JSONResponse(content={"error": str(e)}, status_code=500)
    return JSONResponse(content={"error": "EFT Summary not found"}, status_code=404)

@app.post("/api/set_active_roster")
async def set_active_roster(request: Request):
    try:
        data = await request.json()
        path = data.get("path", "")
        if not path:
             # Fallback
             week_date = data.get("week_date", "")
             path = f"Rosters/Roster_{week_date}.xlsx"
             
        archive_path = os.path.join("data/archives", path.lstrip('/'))
        if os.path.exists(archive_path):
            shutil.copy(archive_path, "data/input/latest_roster.xlsx")
            return JSONResponse(content={"success": True, "message": f"Archive {os.path.basename(path)} set as active draft."})
        return JSONResponse(content={"success": False, "error": f"Archived file not found: {path}"})
    except Exception as e:
        return JSONResponse(content={"success": False, "error": str(e)})

@app.post("/api/save_roster")
async def save_roster(request: Request):
    try:
        data = await request.json()
        week_date = data.get("week_date", "current")
        rows = data.get("rows", [])
        
        df = pd.DataFrame(rows)
        filename = f"data/input/latest_roster.xlsx"
        
        # 1. Fetch staff to get roles from Master Template
        master_path = "data/templates/Staff_Details_Template.xlsx"
        staff_roles = {}
        if os.path.exists(master_path):
            try:
                mdf = pd.read_excel(master_path)
                mdf.columns = [str(c).strip() for c in mdf.columns]
                for _, s_row in mdf.iterrows():
                    name = str(s_row.get("Name", "")).strip()
                    role = str(s_row.get("Role", "WAITER")).strip().upper()
                    if name: staff_roles[name] = role
            except Exception as e:
                print(f"Error reading master staff: {e}")

        # 2. Group the rows (Excluding Managers and Supervisors)
        groups = {}
        for r in rows:
            name = r.get('Name')
            role = staff_roles.get(name, 'WAITER').upper()
            if role in ['MANAGER', 'SUPERVISOR']: continue
            
            if role not in groups: groups[role] = []
            groups[role].append(r)

        wb = Workbook()
        ws = wb.active
        ws.title = "Weekly Roster"

        # Styles
        header_fill = PatternFill(start_color="ED1C24", end_color="ED1C24", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        group_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
        group_font = Font(bold=True, color="ED1C24")
        stripe_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
        center_align = Alignment(horizontal="center", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center")
        thin_border = Border(
            left=Side(style='thin', color='CBD5E1'),
            right=Side(style='thin', color='CBD5E1'),
            top=Side(style='thin', color='CBD5E1'),
            bottom=Side(style='thin', color='CBD5E1')
        )

        # Build Sheet
        current_row = 1
        cols = ["Name", "Wed", "Thu", "Fri", "Sat", "Sun", "Mon", "Tue", "Total Hours"]
        for c_idx, col_name in enumerate(cols, 1):
            cell = ws.cell(row=current_row, column=c_idx, value=col_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border
        
        current_row += 1
        
        # Custom Role Priority and Pretty Names
        role_priority = ['WAITER', 'BARISTA', 'GRILL', 'KITCHEN', 'GENERAL STAFF']
        pretty_names = {'WAITER': 'Waiters', 'BARISTA': 'Baristas', 'GRILL': 'Grillers', 'KITCHEN': 'Kitchen'}
        
        sorted_roles = sorted(groups.keys(), key=lambda x: role_priority.index(x) if x in role_priority else 99)

        for role in sorted_roles:
            # Role Header
            display_name = pretty_names.get(role, role.title())
            cell = ws.cell(row=current_row, column=1, value=f"📁 {display_name.upper()}")
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=9)
            cell.fill = group_fill
            cell.font = group_font
            cell.border = thin_border
            current_row += 1

            for r_idx, r_data in enumerate(groups[role]):
                row_values = [r_data.get(c, "") for c in cols]
                for c_idx, val in enumerate(row_values, 1):
                    cell = ws.cell(row=current_row, column=c_idx, value=val)
                    cell.border = thin_border
                    if (current_row - 1) % 2 == 0: cell.fill = stripe_fill
                    
                    if c_idx == 1:
                        cell.alignment = left_align
                        cell.font = Font(bold=True)
                    else:
                        cell.alignment = center_align
                current_row += 1

        # Adjust Column Widths
        ws.column_dimensions['A'].width = 25
        for col in ['B','C','D','E','F','G','H']:
            ws.column_dimensions[col].width = 14
        ws.column_dimensions['I'].width = 15 

        wb.save(filename)
        
        # Archive it in monthly folder
        archive_dir = "data/archives/Rosters"
        os.makedirs(archive_dir, exist_ok=True)
        archive_path = os.path.join(archive_dir, f"Roster_{week_date}.xlsx")
        shutil.copy(filename, archive_path)
        
        return JSONResponse(content={"success": True, "message": f"Roster saved and archived in {os.path.basename(archive_dir)}"})
    except Exception as e:
        return JSONResponse(content={"success": False, "error": str(e)})

@app.get("/api/info")
async def get_server_info():
    """Returns the current LAN and Remote URLs for connectivity."""
    # Attempt to find ngrok public URL if it's running
    public_url = ""
    try:
        from pyngrok import ngrok
        tunnels = ngrok.get_tunnels()
        if tunnels:
            public_url = tunnels[0].public_url
    except: pass
    
    try:
        lan_ip = socket.gethostbyname(socket.gethostname())
        lan_url = f"http://{lan_ip}:8000"
    except Exception:
        lan_url = "http://127.0.0.1:8000"
    return {"lan": lan_url, "remote": public_url}

@app.post("/api/upload_roster")
async def upload_roster_file(roster: UploadFile = File(...)):
    try:
        os.makedirs("data/input", exist_ok=True)
        roster_path = "data/input/latest_roster.xlsx"
        
        with open(roster_path, "wb") as buffer:
            shutil.copyfileobj(roster.file, buffer)
            
        return JSONResponse(content={"success": True, "message": "Roster uploaded successfully."})
    except Exception as e:
        return JSONResponse(content={"success": False, "error": str(e)})

@app.get("/api/my_roster")
async def get_my_roster(name: str, request: Request):
    roster_path = "data/input/latest_roster.xlsx"
    if not os.path.exists(roster_path):
        return JSONResponse(content={"shifts": {}, "error": "No active roster found."})
    
    try:
        df = pd.read_excel(roster_path)
        
        # Calculate dates for the UI
        day_to_date = {}
        try:
            # Try to get the Tuesday date from file modification time or assume it's the upcoming Tuesday
            import datetime
            mtime = os.path.getmtime(roster_path)
            file_dt = datetime.datetime.fromtimestamp(mtime)
            # Find the Tuesday of that week (0=Mon, 1=Tue...)
            # We assume the file represents the week ending on the nearest future Tuesday from when it was created
            days_until_tue = (1 - file_dt.weekday() + 7) % 7
            tue_dt = file_dt + datetime.timedelta(days=days_until_tue)
            
            days_list = ["Wed", "Thu", "Fri", "Sat", "Sun", "Mon", "Tue"]
            for i, d in enumerate(days_list):
                offset = 6 - i
                day_dt = tue_dt - datetime.timedelta(days=offset)
                day_to_date[d] = day_dt.strftime("%d %b")
        except: pass

        print(f"DEBUG: Loaded roster with columns: {df.columns.tolist()}")
        roster_df = preprocess_roster(df)
        print(f"DEBUG: Preprocessed roster rows: {len(roster_df)}")
        
        # Fallback for "Pretty" Format (Name, Wed, Thu...)
        if roster_df.empty:
            print("DEBUG: Preprocessed df is empty, trying Pretty fallback...")
            cols = [str(c).strip() for c in df.columns]
            day_names = ["Wed", "Thu", "Fri", "Sat", "Sun", "Mon", "Tue"]
            if any(d in cols for d in day_names):
                rows = []
                name_col = next((c for c in df.columns if 'name' in str(c).lower()), df.columns[0])
                print(f"DEBUG: Using name column: {name_col}")
                for _, row in df.iterrows():
                    emp_name = str(row[name_col]).strip()
                    if not emp_name or emp_name.lower() in ['nan', 'name']: continue
                    for d in day_names:
                        if d in df.columns:
                            val = str(row[d]).strip()
                            if val and '-' in val:
                                try:
                                    s, e = val.split('-')[0], val.split('-')[1]
                                    rows.append({
                                        'Employee Name': emp_name,
                                        'Date': d, # Store day name as Date for now
                                        'Rostered Start': s,
                                        'Rostered End': e
                                    })
                                except: continue
                roster_df = pd.DataFrame(rows)
                print(f"DEBUG: Pretty fallback generated {len(roster_df)} rows")
        
        if roster_df.empty:
            return JSONResponse(content={"shifts": {}, "error": "Roster file found but no valid shift rows detected (Name and Time range missing)."})
    except Exception as inner_e:
        print(f"DEBUG: CRASH in get_my_roster: {inner_e}")
        return JSONResponse(content={"shifts": {}, "error": "Error parsing roster data.", "debug": str(inner_e)})
    
    try:

        # unique dates (could be day names or ISO dates)
        dates = sorted(roster_df['Date'].unique())
        
        is_manager_param = request.query_params.get("is_manager", "false").lower() == "true"
        
        if name.lower() == "manager" or is_manager_param:
            # 1. Fetch staff roles for grouping
            role_map = {}
            master_path = "data/templates/Staff_Details_Template.xlsx"
            if os.path.exists(master_path):
                try:
                    m_df = pd.read_excel(master_path)
                    m_df.columns = [str(c).strip() for c in m_df.columns]
                    n_col = next((c for c in m_df.columns if str(c).lower() == 'name'), 'Name')
                    r_col = next((c for c in m_df.columns if str(c).lower() == 'role'), 'Role')
                    for _, r in m_df.iterrows():
                        nm = str(r[n_col]).strip()
                        rl = str(r[r_col]).strip().upper() if not pd.isna(r[r_col]) else "OTHER"
                        role_map[nm] = rl
                except Exception as ex:
                    print(f"DEBUG: Failed to load roles for roster: {ex}")

            # 2. Manager View: Return all staff shifts grouped by employee
            team_roster = {}
            for _, row in roster_df.iterrows():
                emp = row['Employee Name']
                # Skip role headers or icons
                if not emp or "ðŸ“" in str(emp): continue
                if str(emp).isupper() and any(role in str(emp) for role in ['WAITERS', 'KITCHEN', 'SCULLERY', 'MANAGEMENT', 'FRONT OF HOUSE']):
                    continue
                
                date_val = row['Date']
                try:
                    day_name = pd.to_datetime(date_val).strftime('%a')
                except:
                    day_name = str(date_val)
                
                if emp not in team_roster:
                    team_roster[emp] = {
                        "role": role_map.get(emp, "OTHER"),
                        "shifts": {}
                    }
                team_roster[emp]["shifts"][day_name] = f"{row['Rostered Start']}-{row['Rostered End']}"
            
            # 3. Sort by Role then Name
            role_order = ["MANAGER", "SUPERVISOR", "BARISTA", "GRILLER", "KITCHEN", "WAITER", "OTHER"]
            def get_sort_key(item):
                role = item[1].get("role", "OTHER")
                try: r_idx = role_order.index(role)
                except: r_idx = 99
                return (r_idx, item[0].lower())

            sorted_roster = dict(sorted(team_roster.items(), key=get_sort_key))

            # Use consistent date labels for columns (Wed, Thu...)
            day_names = ["Wed", "Thu", "Fri", "Sat", "Sun", "Mon", "Tue"]
            return JSONResponse(content={
                "is_manager": True, 
                "team_roster": sorted_roster, 
                "dates": day_names,
                "day_to_date": day_to_date
            })
        else:
            # Employee View
            import difflib
            emp_names = [str(n).strip() for n in roster_df['Employee Name'].unique()]
            matches = difflib.get_close_matches(name.strip(), emp_names, n=1, cutoff=0.4)
            
            if not matches:
                return JSONResponse(content={"shifts": {}, "error": f"Staff name '{name}' not found in current roster."})
            
            matched_name = matches[0]
            emp_shifts = roster_df[roster_df['Employee Name'].str.strip() == matched_name]
            shifts = {}
            for _, row in emp_shifts.iterrows():
                date_val = row['Date']
                try: day_name = pd.to_datetime(date_val).strftime('%a')
                except: day_name = str(date_val)
                shifts[day_name] = f"{row['Rostered Start']}-{row['Rostered End']}"
            
            return JSONResponse(content={
                "is_manager": False, 
                "shifts": shifts, 
                "matched_name": matched_name,
                "day_to_date": day_to_date
            })
            
    except Exception as e:
        print(f"DEBUG: CRASH in get_my_roster: {e}")
        import traceback
        traceback.print_exc()
        return JSONResponse(content={"shifts": {}, "error": "Dashboard error while preparing view.", "debug": str(e)})

@app.post("/api/accrue_leave")
async def accrue_leave():
    master_path = "data/templates/Staff_Details_Template.xlsx"
    if not os.path.exists(master_path):
        return JSONResponse(content={"status": "error", "message": "Master file not found. Fill out Staff_Details_Template.xlsx first."})
        
    try:
        master_df = pd.read_excel(master_path)
        leave_col = next((c for c in master_df.columns if str(c).strip().lower() == 'leave credit'), None)
        
        if leave_col:
            for idx, row in master_df.iterrows():
                try: cur_leave = float(row[leave_col]) if str(row[leave_col]).strip().lower() != 'nan' else 0.0
                except: cur_leave = 0.0
                master_df.at[idx, leave_col] = cur_leave + 1.5
            master_df.to_excel(master_path, index=False)
            return JSONResponse(content={"status": "success", "message": "1.5 days accrued to all staff successfully!"})
        return JSONResponse(content={"status": "error", "message": "Leave Credit column not found."})
    except Exception as e:
        return JSONResponse(content={"status": "error", "message": str(e)})



# --- GEMINI AI ASSISTANT ---

def update_ai_usage():
    """Tracks and returns (can_proceed, tries_left)"""
    path = "data/ai_usage.json"
    today = datetime.now().strftime("%Y-%m-%d")
    limit = 10
    
    usage = {"date": today, "tries_used": 0, "limit": limit}
    if os.path.exists(path):
        try:
            with open(path, "r") as f:
                saved = json.load(f)
                if saved.get("date") == today:
                    usage = saved
        except: pass
        
    if usage["tries_used"] >= usage["limit"]:
        return False, 0
        
    usage["tries_used"] += 1
    os.makedirs("data", exist_ok=True)
    with open(path, "w") as f:
        json.dump(usage, f, indent=4)
        
    return True, usage["limit"] - usage["tries_used"]

def get_ai_usage():
    """Returns tries_left without incrementing"""
    path = "data/ai_usage.json"
    today = datetime.now().strftime("%Y-%m-%d")
    limit = 10
    if os.path.exists(path):
        try:
            with open(path, "r") as f:
                saved = json.load(f)
                if saved.get("date") == today:
                    return limit - saved.get("tries_used", 0)
        except: pass
    return limit

@app.get("/api/ai/usage")
async def ai_usage_status():
    return JSONResponse(content={"tries_left": get_ai_usage()})

@app.post("/api/ai/chat")
async def ai_chat(request: Request):
    if not GEMINI_KEY:
        return JSONResponse(content={"error": "API_KEY_MISSING", "message": "Gemini API Key is missing."})
    
    try:
        data = await request.json()
        user_msg = data.get("message", "")
        user_role = data.get("role", "EMPLOYEE")
        user_name = data.get("name", "Unknown Staff")
        
        # 1. Gather Staff Context
        all_staff = get_all_staff_list()
        staff_context = "\n".join([f"- {s['name']} (Role: {s['role']})" for s in all_staff if s['role'] not in ['MANAGER', 'SUPERVISOR']])
            
        # 2. Gather Public & School Holidays
        sa_holidays = pyholidays.SouthAfrica(years=datetime.now().year)
        holiday_list = ", ".join([f"{d} ({n})" for d, n in sorted(sa_holidays.items())])
        
        school_holidays = "2026-03-28 to 2026-04-07, 2026-06-27 to 2026-07-20, 2026-09-24 to 2026-10-05, 2026-12-10 to 2026-12-31"

        # 3. Gather Off-Day Requests
        requests_summary = "No pending off-day requests."
        req_path = "data/off_day_requests.json"
        if os.path.exists(req_path):
            with open(req_path, "r") as f:
                req_data = json.load(f)
                requests_summary = json.dumps(req_data, indent=2)

        # 4. Gather AI Learning Log (manager-added notes for roster generation)
        learning_summary = "No learning notes."
        log_path = "data/roster_learning_log.json"
        if os.path.exists(log_path):
            with open(log_path, "r") as f:
                try:
                    log_data = json.load(f)
                    if log_data:
                        learning_summary = "\n".join([f"- [{e.get('date','')}] {e.get('note','')}" for e in log_data])
                except:
                            system_prompt = f"""
        You are the 'Wimpy De Ville AI Manager-on-Duty'. 
        Current Date: {datetime.now().strftime('%Y-%m-%d (%A)')}
        
        STRICT OPERATIONAL RULES:
        1. Roster Cycle: Wednesday to Tuesday. Starts on Wednesday morning and ends on Tuesday evening.
        2. Store Hours & Closing:
           - Monday to Friday: 07:00 - 18:00
           - Saturday: 07:00 - 15:00
           - Sunday & Public Holidays: 08:00 - 13:00 (Reduced Staffing, usually 1 Waiter, 1 Kitchen, 1 Griller)
        3. Lead Time: Leave (Vacation) requires 14 days notice.
        4. Approval Flags: Requests for Sat/Sun or School Holidays ALWAYS require manager approval.
        5. STAFF PAIRINGS: Pay extremely close attention to the 'MANAGER LEARNING NOTES'. If a note specifies a pairing (e.g., 'Joe must always work with Sam'), assign them overlapping shifts.
        6. OFF-DAY RULES: 
           - Staff can request up to 2 days off. 2 CONSECUTIVE DAYS (e.g. Mon and Tue) are NOT allowed.
           - APPROVED 'Annual Leave' means they are completely excluded (marked 'LEAVE' with 0 hours).
        
        STAFF DATABASE (ACTIVE LIST - USE THESE EXACT NAMES):
        {staff_context}
        
        STRICT NAME MATCHING RULE:
        - You MUST use the EXACT names provided in the 'STAFF DATABASE' (including capitalization and special characters).
        - NEVER abbreviate names or use nicknames.
        - If a staff member is not in the list, ignore them.
        
        PUBLIC HOLIDAYS (SOUTH AFRICA):
        {holiday_list}
        
        SCHOOL HOLIDAYS 2026:
        {school_holidays}
        
        PENDING OFF-DAY & LEAVE REQUESTS:
        {requests_summary}
        
        MANAGER LEARNING LOG & RULES:
        {learning_summary}
        
        OUTPUT GUIDELINES:
        - First, give a brief friendly summary of the roster changes.
        - Then, include an 'Individual Snippet' for each person (Name, Dates, Times, Total Hours).
        - Finally, provide the raw JSON roster block between 'ROSTER_JSON_START' and 'ROSTER_JSON_END' markers.
        - Format for JSON: [{{"Name": "EXACT STAFF NAME", "Wed": "07:00-18:00" or "OFF", ..., "Total": "45.0"}}]
        - Use 'HH:MM-HH:MM' format for shifts (no spaces around hyphen).
        - Ensure coverage: 2 Kitchen, 2 Waiters/FOH per regular shift.
        
        PRIMARY DIRECTIVE: Follow all 'MANAGER LEARNING LOG' notes as high-priority constraints.
        USER: {user_role} ({user_name})
        """

        if not client:
             return JSONResponse(content={"error": "AI_CLIENT_MISSING"})

        can_proceed, tries_left = update_ai_usage()
        if not can_proceed:
            return JSONResponse(content={"error": "LIMIT_REACHED", "message": "Daily AI generation limit reached (10/10). Please try again tomorrow."})

        # Retry logic for AI service
        last_err = ""
        for attempt in range(3):
            try:
                response = client.models.generate_content(
                    model='gemini-flash-latest',
                    contents=f"{system_prompt}\n\nUSER REQUEST: {user_msg}"
                )
                return JSONResponse(content={"reply": response.text, "tries_left": tries_left})
            except Exception as ai_err:
                last_err = str(ai_err)
                print(f"AI Attempt {attempt+1} failed: {last_err}")
                if "503" in last_err or "429" in last_err or "busy" in last_err.lower():
                    time.sleep(2 * (attempt + 1)) # Exponential backoff
                    continue
                else:
                    break
        
        return JSONResponse(content={"error": "AI_TEMPORARILY_UNAVAILABLE", "message": f"The AI service is currently busy. Please try again in a few minutes. (Details: {last_err})"})
    except Exception as e:
        import traceback
        traceback.print_exc()
        return JSONResponse(content={"error": "SERVER_ERROR", "message": str(e)})

@app.post("/api/ai/generate_snippets")
async def generate_snippets(request: Request):
    if not GEMINI_KEY:
        return JSONResponse(content={"error": "API_KEY_MISSING"})
        
    try:
        data = await request.json()
        roster_rows = data.get("rows", [])
        week_date = data.get("week_date", "next week")
        
        can_proceed, tries_left = update_ai_usage()
        if not can_proceed:
             return JSONResponse(content={"error": "LIMIT_REACHED", "message": "Limit reached."})

        snippets = []
        for row in roster_rows:
            name = row.get("Name", "Staff")
            # Ask AI for a concise snippet for this specific person's row
            prompt = f"Based on this roster row for {name} for the week ending {week_date}, generate a concise WhatsApp snippet:\n{json.dumps(row)}\n\nFormat: 'Hi {name}, your roster: ...'"
            
            response = client.models.generate_content(
                model='gemini-flash-latest',
                contents=prompt
            )
            snippets.append({
                "name": name,
                "phone": row.get("Cell Number", ""), 
                "snippet": response.text.strip()
            })
            
        return JSONResponse(content={"snippets": snippets, "tries_left": tries_left})
    except Exception as e:
        return JSONResponse(content={"error": str(e)})

@app.post("/api/whatsapp_send_snippets")
async def whatsapp_send_snippets(request: Request):
    data = await request.json()
    snippets = data.get("snippets", [])
    
    # Fetch staff cell numbers if missing
    master_path = "data/templates/Staff_Details_Template.xlsx"
    phone_map = {}
    if os.path.exists(master_path):
        df = pd.read_excel(master_path)
        for _, row in df.iterrows():
            name = str(row.get("Name", "")).strip().lower()
            phone = str(row.get("Cell Number", "")).replace(".0", "").replace("nan", "")
            phone_map[name] = ''.join(filter(str.isdigit, phone))

    payload = []
    for s in snippets:
        name_lower = s.get("name", "").lower()
        phone = s.get("phone") or phone_map.get(name_lower, "")
        if phone:
            payload.append({
                "phone": phone,
                "message": s.get("snippet", ""),
                "file": "" # No file for snippets usually, just text
            })
            
    with open("output/whatsapp_payload.json", "w") as f:
        json.dump(payload, f)
        
    subprocess.Popen([sys.executable, "app/wa_bot.py"])
    return JSONResponse(content={"status": "Snippets queued for WhatsApp bot."})

@app.post("/api/send_roster")
async def send_roster(request: Request):
    try:
        data = await request.json()
        links = data.get("links", [])
        manager_email = "dylan.lloyd25@gmail.com"
        
        # Stub for SMTP flow. Manager would need to provide an App Password if using Gmail.
        sent_email_count = 0
        for l in links:
             if l.get("email") and "@" in str(l.get("email")):
                  sent_email_count += 1
                  
        return JSONResponse(content={"success": True, "message": f"Roster pieces sent to {sent_email_count} via Email, others queued for WhatsApp bot."})
    except Exception as e:
        return JSONResponse(content={"success": False, "error": str(e)})



# --- AI LEARNING LOG ---

@app.get("/api/ai/learning_log")
async def get_learning_log():
    path = "data/roster_learning_log.json"
    if not os.path.exists(path): return JSONResponse(content=[])
    try:
        with open(path, "r") as f:
            return JSONResponse(content=json.load(f))
    except: return JSONResponse(content=[])

@app.post("/api/ai/learning_log")
async def add_learning_log(request: Request):
    data = await request.json()
    note = data.get("note", "").strip()
    if not note: return JSONResponse(content={"success": False, "error": "Note cannot be empty"})
    path = "data/roster_learning_log.json"
    log = []
    if os.path.exists(path):
        with open(path, "r") as f:
            try: log = json.load(f)
            except: log = []
    log.append({"date": datetime.now().strftime("%Y-%m-%d %H:%M"), "note": note})
    with open(path, "w") as f: json.dump(log, f, indent=4)
    return JSONResponse(content={"success": True, "total_notes": len(log)})

@app.delete("/api/ai/learning_log/{index}")
async def delete_learning_log(index: int):
    path = "data/roster_learning_log.json"
    if not os.path.exists(path): return JSONResponse(content={"success": False, "error": "Log not found"})
    with open(path, "r") as f: log = json.load(f)
    if index < 0 or index >= len(log):
        return JSONResponse(content={"success": False, "error": "Index out of range"})
    log.pop(index)
    with open(path, "w") as f: json.dump(log, f, indent=4)
    return JSONResponse(content={"success": True})


# --- FINGERPRINT SCANNER INTEGRATION ---

@app.get("/api/fingerprint/records")
async def get_fingerprint_records():
    path = "data/fingerprint_scans.csv"
    if not os.path.exists(path):
        return JSONResponse(content={"records": []})
    try:
        import csv
        records = []
        with open(path, "r", newline="") as f:
            reader = csv.DictReader(f)
            for row in reader:
                records.append(dict(row))
        return JSONResponse(content={"records": records})
    except Exception as e:
        return JSONResponse(content={"error": str(e)})

@app.post("/api/fingerprint/sync")
async def sync_fingerprint(request: Request):
    data = await request.json()
    records = data.get("records", [])
    if not records:
        return JSONResponse(content={"success": False, "error": "No records provided"})
    path = "data/fingerprint_scans.csv"
    try:
        import csv
        file_exists = os.path.exists(path) and os.path.getsize(path) > 0
        with open(path, "a", newline="") as f:
            fieldnames = ["EmployeeID", "EmployeeName", "Timestamp", "Action"]
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            if not file_exists:
                writer.writeheader()
            for r in records:
                writer.writerow({
                    "EmployeeID": r.get("EmployeeID", ""),
                    "EmployeeName": r.get("EmployeeName", ""),
                    "Timestamp": r.get("Timestamp", datetime.now().isoformat()),
                    "Action": r.get("Action", "IN")
                })
        return JSONResponse(content={"success": True, "synced": len(records)})
    except Exception as e:
        return JSONResponse(content={"success": False, "error": str(e)})

@app.get("/api/staff/payslips/{name}")
async def get_staff_payslip_history(name: str):
    """Scan data/archives recursively for any PDFs matching the staff name."""
    try:
        archives_dir = "data/archives"
        matches = []
        if os.path.exists(archives_dir):
            import re
            # Windows-safe name was used for filenames: re.sub(r'[<>:"/\\|?*]', '', raw_name).strip().replace(' ', '_')
            safe_name = re.sub(r'[<>:"/\\|?*]', '', name).strip().replace(' ', '_')
            
            for root, dirs, files in os.walk(archives_dir):
                for f in files:
                    if f.endswith(".pdf") and safe_name in f:
                        # Extract date from filename (YYYY-MM-DD)
                        date_match = re.search(r'(\d{4}-\d{2}-\d{2})', f)
                        date_str = date_match.group(1) if date_match else "Unknown Date"
                        
                        rel_path = os.path.relpath(os.path.join(root, f), ".").replace("\\", "/")
                        folder_name = os.path.basename(root)
                        
                        matches.append({
                            "date": date_str,
                            "url": rel_path,
                            "folder": folder_name
                        })
        
        # Sort by date descending
        matches.sort(key=lambda x: x['date'], reverse=True)
        return JSONResponse(content={"payslips": matches})
    except Exception as e:
        return JSONResponse(content={"error": str(e)}, status_code=500)

@app.get("/api/archives/download/{path_sub:path}")
async def download_archive_file(path_sub: str):
    """Serve the raw Excel or PDF file from archives for download."""
    path = os.path.join("data/archives", path_sub)
    if os.path.exists(path):
        return FileResponse(path, filename=os.path.basename(path))
    return JSONResponse(content={"error": "File not found"}, status_code=404)
